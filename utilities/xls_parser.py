import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    workbook.close()
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    workbook.close()
    return sheet.max_column


def read_data(file, sheet_name, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    workbook.close()
    return sheet.cell(row=rownum, column=columnnum).value


def write_data(file, sheet_name, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)
    workbook.close()
